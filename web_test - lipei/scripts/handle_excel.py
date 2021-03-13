# -*- coding: utf-8 -*-
"""
    Create Time: 2019/6/26 11:13
    Author: 作者
"""

from openpyxl import load_workbook
from scripts.handel_config import config
from scripts.constants import TEST_CASES_PATH
class HandleExcel:

    def __init__(self,filename,sheetname=None):
        self.filename , self.sheetname = filename, sheetname

    def get_cases(self):
        #1.打开excel文件
        wb = load_workbook(self.filename)
        # 2.定位表单
        # 如果你有传sheetname, 那么就获取指定的表单
        # 如果你没有传sheetname, 那么就获取第一个表单
        if self.sheetname is None:
            ws = wb.active
        else:
            ws = wb[self.sheetname]
        #获取所有用例数据
        total_data= []
        values = tuple(ws.iter_rows(min_row=1,max_row=1,values_only=True))
        sheet_head = values[0]

        for data in tuple(ws.iter_rows(min_row=2,values_only=True)):
            total_data.append(dict(zip(sheet_head,data)))
        return total_data

    def write_result(self,row,actual,result):
        wb = load_workbook(self.filename)
        if self.sheetname is None:
            ws = wb.active
        else:
            ws = wb[self.sheetname]

        if isinstance(row,int) and (2<= row <= ws.max_row):
            ws.cell(row=row,column=config.get_int('excel','actual_col'), value=actual)
            ws.cell(row=row,column=config.get_int('excel','result_col'), value=result)
            wb.save(self.filename)
        else:
            print('传入的行号有误，行号应大于1的整数')

if __name__ == '__main__':
    excel = HandleExcel(TEST_CASES_PATH,'Sheet1')
    result = excel.get_cases()
    print(result)
    print(len(result))



